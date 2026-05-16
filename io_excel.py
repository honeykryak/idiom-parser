"""
Модуль ввода-вывода Excel — чтение и запись предложений через openpyxl.
"""

from openpyxl import Workbook, load_workbook

# Название столбца, из которого читаются контексты
FULL_CONTEXT_COLUMN = "Full context"


def read_sentences(path: str) -> list[str]:
    """Читает предложения из столбца 'Full context' первого листа.

    Пропускает пустые ячейки и убирает пробелы по краям.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        col_index = list(header_row).index(FULL_CONTEXT_COLUMN)
    except ValueError:
        wb.close()
        raise ValueError(
            f"Столбец '{FULL_CONTEXT_COLUMN}' не найден. "
            f"Доступные столбцы: {list(header_row)}"
        )

    sentences: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col_index]
        if value is not None:
            text = str(value).strip()
            if text:
                sentences.append(text)
    wb.close()
    return sentences


def write_sentences(sentences: list[str], path: str) -> None:
    """Записывает список предложений в столбец A нового файла."""
    wb = Workbook()
    ws = wb.active
    for sent in sentences:
        ws.append([sent])
    wb.save(path)
    wb.close()
